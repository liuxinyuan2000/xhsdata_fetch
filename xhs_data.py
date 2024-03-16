from DrissionPage import ChromiumPage
import pandas as pd
from tqdm import tqdm
import time
from urllib.parse import quote
import random
import re
import openpyxl


def sign_in():
    sign_in_page = ChromiumPage()
    sign_in_page.get('https://www.xiaohongshu.com')
    print("请扫码登录")
    # 第一次运行需要扫码登录
    time.sleep(20)


def search(keyword):
    global page
    page = ChromiumPage()
    page.get(f'https://www.xiaohongshu.com/search_result?keyword={keyword}&source=web_search_result_notes')


def select_category(category):
    # 切换搜索结果到仅包含视频
    # 定位包含笔记信息的sections
    content_container = page.ele('.content-container')
    videos = content_container.ele(f'text={category}')
    videos.click()


def get_info():
    print(f"第{i}次爬取")
    # 定位包含笔记信息的sections
    container = page.ele('.feeds-page')
    sections = container.eles('.note-item')
    # 提取笔记信息
    for section in sections:
        try:
            # 文章链接
            note_link = section.ele('tag:a', timeout=0).link
            # 标题、作者、点赞
            footer = section.ele('.footer', timeout=0)
            # 标题
            try:
                title = footer.ele('.title', timeout=0).text
            except:
                title = "标题为空"

            # 作者
            author_wrapper = footer.ele('.author-wrapper')
            author = author_wrapper.ele('.author').text
            author_link = author_wrapper.ele('tag:a', timeout=0).link
            author_img = author_wrapper.ele('tag:img', timeout=0).link

            # 点赞
            like = footer.ele('.like-wrapper like-active').text
            # 处理以w结尾
            if like.endswith("w"):
                # 使用正则表达式查找数字
                numbers = re.findall(r'\d+\.\d+|\d+', like)
                # 将数字字符串转换为浮点数
                float_value = float(numbers[0])
                # 将浮点数乘以 10000 以表示万
                like = int(float_value * 10000)
            else:
                like = int(like)

            print(title, author, like, note_link, author_link, author_img)
            contents.append([title, author, note_link, author_link, author_img, like])

        except:
            pass
    print(f"爬取{i}次，总计获取{len(contents)}条")


def page_scroll_down():
    print("********下滑页面********")
    # 生成一个随机时间
    random_time = random.uniform(0.5, 1.5)
    # 暂停
    time.sleep(random_time)
    # time.sleep(1)
    # page.scroll.down(5000)
    page.scroll.to_bottom()


def crawler(times):
    global i
    for i in tqdm(range(1, times + 1)):
        get_info()
        page_scroll_down()


def auto_resize_column(excel_path):
    """自适应列宽度"""
    wb = openpyxl.load_workbook(excel_path)
    worksheet = wb.active
    # 循环遍历工作表中的1-2列
    for col in worksheet.iter_cols(min_col=1, max_col=2):
        max_length = 0
        # 列名称
        column = col[0].column_letter
        # 循环遍历列中的所有单元格
        for cell in col:
            try:
                # 如果当前单元格的值长度大于max_length，则更新 max_length 的值
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # 计算调整后的列宽度
        adjusted_width = (max_length + 2) * 2
        # 使用 worksheet.column_dimensions 属性设置列宽度
        worksheet.column_dimensions[column].width = adjusted_width

        # 循环遍历工作表中的3-5列
        for col in worksheet.iter_cols(min_col=3, max_col=5):
            max_length = 0
            column = col[0].column_letter  # Get the column name

            # 使用 worksheet.column_dimensions 属性设置列宽度
            worksheet.column_dimensions[column].width = 15

    wb.save(excel_path)


def save_to_excel(data):
    # 保存到excel文件
    name = ['标题', '作者', '笔记链接', '作者主页', '作者头像', '点赞数']
    df = pd.DataFrame(columns=name, data=data)

    # # 写入原文件前清除openpyxl不支持的字符
    # for col in df[['title', 'author']]:
    #     df[col] = df[col].apply(lambda x: ILLEGAL_CHARACTERS_RE.sub(r'', str(x) if not pd.isna(x) else ''))

    df['点赞数'] = df['点赞数'].astype(int)
    # 删除重复行
    df = df.drop_duplicates()
    # 按点赞 降序排序
    df = df.sort_values(by='点赞数', ascending=False)
    # 文件路径
    # global excel_path
    excel_path = f"小红书搜索结果-{category}-{keyword}-{df.shape[0]}条.xlsx"
    df.to_excel(excel_path, index=False)
    print(f"总计向下翻页{times}次，获取到{len(data)}条，去重后剩余{df.shape[0]}条")
    print(f"数据已保存到：{excel_path}")

    # 自动调整excel表格列宽
    auto_resize_column(excel_path)
    print("已完成自动调整excel表格列宽")


if __name__ == '__main__':
    # contents列表用来存放所有爬取到的信息
    contents = []

    # 1、设置搜索关键词
    keyword = "互粉"
    # 2、设置向下翻页爬取次数
    times = 20
    # 3、设置选择笔记类型，可选择 “全部、图文、视频”，默认是“全部”
    category = "全部"

    # 第1次运行需要登录，后面不用登录，可以注释掉
    # sign_in()

    # 关键词转为 url 编码
    keyword_temp_code = quote(keyword.encode('utf-8'))
    keyword_encode = quote(keyword_temp_code.encode('gb2312'))

    # 根据关键词搜索小红书文章
    search(keyword_encode)

    select_category(category)

    # 根据设置的次数，开始爬取数据
    crawler(times)

    # 爬到的数据保存到本地excel文件
    save_to_excel(contents)
